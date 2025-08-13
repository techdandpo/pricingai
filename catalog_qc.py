import io
import streamlit as st
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill

# Default column suggestions
DEFAULT_SKU_COL = 'Dandpo SKU'
DEFAULT_QTY_COL = 'Quantity'


def _clean_df(df: pd.DataFrame) -> pd.DataFrame:
    df = df.dropna(how='all')
    df = df.fillna('')
    # Try convert Quantity to numeric safely
    if 'Quantity' in df.columns:
        df['Quantity'] = pd.to_numeric(df['Quantity'], errors='coerce').fillna(0)
    return df


def _suggest_price_column(columns: list[str]) -> str | None:
    # Prefer columns containing 'Customer Price' then 'Price'
    for pref in ['Customer Price', 'Price']:
        for col in columns:
            if pref.lower() in col.lower():
                return col
    # Fallback: None
    return None


def _suggest_column(columns: list[str], default_name: str) -> str | None:
    # First try exact match
    if default_name in columns:
        return default_name
    # Then try case-insensitive match
    for col in columns:
        if default_name.lower() in col.lower():
            return col
    # Fallback: None
    return None


def _build_product_key(df: pd.DataFrame, sku_col: str, qty_col: str) -> pd.Series:
    key_cols = [sku_col, qty_col]
    return df[key_cols].astype(str).agg(' | '.join, axis=1)


def _styled_preview(qc_df: pd.DataFrame):
    def color_row(row):
        status = row.get('QC Status', '')
        color = ''
        if status == 'MATCH':
            color = 'background-color: #90EE90'  # green
        elif status == 'MISMATCH':
            color = 'background-color: #FF7F7F'  # red
        elif status == 'MISSING':
            color = 'background-color: #FFF59D'  # yellow
        return [color] * len(row)

    return qc_df.style.apply(color_row, axis=1)


def _excel_with_colors(qc_df: pd.DataFrame) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = 'Catalog QC'

    green = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
    red = PatternFill(start_color='FF7F7F', end_color='FF7F7F', fill_type='solid')
    yellow = PatternFill(start_color='FFF59D', end_color='FFF59D', fill_type='solid')

    # Write header
    ws.append(list(qc_df.columns))

    status_idx = qc_df.columns.get_loc('QC Status')

    for _, row in qc_df.iterrows():
        values = [row[col] for col in qc_df.columns]
        ws.append(values)
        # Apply row color by status
        status = str(row['QC Status'])
        fill = None
        if status == 'MATCH':
            fill = green
        elif status == 'MISMATCH':
            fill = red
        elif status == 'MISSING':
            fill = yellow
        if fill is not None:
            r = ws.max_row
            for c in range(1, len(qc_df.columns) + 1):
                ws.cell(row=r, column=c).fill = fill

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio


def catalog_qc():
    st.header('🧪 Catalog QC')

    col_left, col_right = st.columns(2)

    with col_left:
        file_a = st.file_uploader('Upload Sheet A (CSV)', type=['csv'], key='qc_a')
    with col_right:
        file_b = st.file_uploader('Upload Sheet B (CSV)', type=['csv'], key='qc_b')

    if not (file_a and file_b):
        st.info('Upload both sheets to start QC.')
        return

    # Read and clean
    try:
        df_a = _clean_df(pd.read_csv(file_a))
        df_b = _clean_df(pd.read_csv(file_b))
    except Exception as e:
        st.error(f'Failed to read files: {e}')
        return

    # Column selection section
    st.markdown('### Select columns for comparison')
    
    # SKU column selection
    st.markdown('**SKU Column (mandatory):**')
    with col_left:
        default_sku_a = _suggest_column(list(df_a.columns), DEFAULT_SKU_COL)
        sku_col_a = st.selectbox('SKU column in Sheet A', options=list(df_a.columns), 
                                index=(list(df_a.columns).index(default_sku_a) if default_sku_a in df_a.columns else 0), 
                                key='sku_a')
    with col_right:
        default_sku_b = _suggest_column(list(df_b.columns), DEFAULT_SKU_COL)
        sku_col_b = st.selectbox('SKU column in Sheet B', options=list(df_b.columns), 
                                index=(list(df_b.columns).index(default_sku_b) if default_sku_b in df_b.columns else 0), 
                                key='sku_b')
    
    # Quantity column selection
    st.markdown('**Quantity Column (mandatory):**')
    with col_left:
        default_qty_a = _suggest_column(list(df_a.columns), DEFAULT_QTY_COL)
        qty_col_a = st.selectbox('Quantity column in Sheet A', options=list(df_a.columns), 
                                index=(list(df_a.columns).index(default_qty_a) if default_qty_a in df_a.columns else 0), 
                                key='qty_a')
    with col_right:
        default_qty_b = _suggest_column(list(df_b.columns), DEFAULT_QTY_COL)
        qty_col_b = st.selectbox('Quantity column in Sheet B', options=list(df_b.columns), 
                                index=(list(df_b.columns).index(default_qty_b) if default_qty_b in df_b.columns else 0), 
                                key='qty_b')

    # Price column selection
    st.markdown('### Select price columns to compare')
    with col_left:
        default_a = _suggest_price_column(list(df_a.columns))
        price_col_a = st.selectbox('Price column in Sheet A', options=list(df_a.columns), index=(list(df_a.columns).index(default_a) if default_a in df_a.columns else 0), key='price_a')
    with col_right:
        default_b = _suggest_price_column(list(df_b.columns))
        price_col_b = st.selectbox('Price column in Sheet B', options=list(df_b.columns), index=(list(df_b.columns).index(default_b) if default_b in df_b.columns else 0), key='price_b')

    # Build keys
    df_a['_QC_KEY'] = _build_product_key(df_a, sku_col_a, qty_col_a)
    df_b['_QC_KEY'] = _build_product_key(df_b, sku_col_b, qty_col_b)

    # Coerce price columns to numeric
    df_a['_QC_PRICE'] = pd.to_numeric(df_a[price_col_a], errors='coerce')
    df_b['_QC_PRICE'] = pd.to_numeric(df_b[price_col_b], errors='coerce')

    # Prepare subset for merge
    subset_a = df_a[['_QC_KEY', '_QC_PRICE']].rename(columns={'_QC_PRICE': 'Price A'})
    subset_b = df_b[['_QC_KEY', '_QC_PRICE']].rename(columns={'_QC_PRICE': 'Price B'})

    merged = pd.merge(subset_a, subset_b, on='_QC_KEY', how='outer')

    # Determine status
    def status_row(row):
        a = row.get('Price A')
        b = row.get('Price B')
        if pd.isna(a) or pd.isna(b):
            return 'MISSING'
        # Compare with 2-decimal rounding for stability
        return 'MATCH' if round(float(a), 2) == round(float(b), 2) else 'MISMATCH'

    merged['QC Status'] = merged.apply(status_row, axis=1)

    # Split key back to columns for readability
    display_keys = [sku_col_a, qty_col_a]
    
    def split_key(s: str) -> list[str]:
        parts = s.split(' | ')
        # pad/truncate
        if len(parts) < len(display_keys):
            parts += [''] * (len(display_keys) - len(parts))
        return parts[:len(display_keys)]

    keys_df = merged['_QC_KEY'].apply(split_key).apply(pd.Series)
    
    # Ensure unique column names to avoid styling conflicts
    unique_display_keys = []
    for i, key in enumerate(display_keys):
        if key in unique_display_keys:
            # If duplicate, add suffix
            unique_display_keys.append(f"{key}_{i}")
        else:
            unique_display_keys.append(key)
    
    keys_df.columns = unique_display_keys

    qc_df = pd.concat([keys_df, merged[['Price A', 'Price B', 'QC Status']]], axis=1)

    st.subheader('QC Preview')
    st.dataframe(_styled_preview(qc_df), use_container_width=True)

    # Downloads
    excel_buffer = _excel_with_colors(qc_df)
    st.download_button(
        '📥 Download QC Report (Excel, colored)',
        excel_buffer,
        file_name='Catalog_QC_Report.xlsx',
        mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
