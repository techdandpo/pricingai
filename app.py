import streamlit as st
import pandas as pd
import io
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

st.set_page_config(page_title="📊 Bidding Sheet Builder", layout="wide")

# Inject custom CSS for brand styling and logo
st.markdown("""
    <style>
        .main {
            background-color: #f8f9fa;
        }
        .stApp {
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            color: #000000;
        }
        .stApp h1, .stApp h2, .stApp h3, .stApp h4 {
            color: #185f5d;
            font-weight: 600;
        }
        .stApp h1 {
            font-size: 2.5rem;
            margin-bottom: 1rem;
        }
        .stApp h2 {
            font-size: 2rem;
        }
        .stApp h3 {
            font-size: 1.5rem;
        }
        header .block-container {
            padding-top: 2rem;
        }
        .reportview-container .main footer {visibility: hidden;}
        .logo-header {
            display: flex;
            align-items: left;
            justify-content: left;
        }
        .title-section {
            text-align: left;
            margin-bottom: 2rem;
        }
        .title-section h1 {
            margin-bottom: 0.5rem;
        }
        .separator-line {
            height: 3px;
            background-color: #185f5d;
            margin: 0 auto;
            width: 100%;
        }
        .stTabs [data-baseweb="tab-list"] {
            gap: 8px;
            padding: 0 1rem;
        }
        .stTabs [data-baseweb="tab"] {
            background-color: #e9ecef;
            border-radius: 8px 8px 0 0;
            color: #000000;
            font-weight: 500;
            padding: 0.75rem 1.5rem;
            min-width: 120px;
            text-align: center;
            white-space: nowrap;
            overflow: hidden;
            text-overflow: ellipsis;
        }
        .stTabs [aria-selected="true"] {
            background-color: #185f5d;
            color: white;
        }
        .stTabs [data-baseweb="tab"] span {
            font-size: 0.9rem;
            line-height: 1.2;
        }
        .stButton > button {
            background-color: #185f5d;
            color: white;
            border: none;
            border-radius: 6px;
            padding: 0.5rem 1rem;
            font-weight: 500;
            transition: all 0.3s ease;
        }
        .stButton > button:hover {
            background-color: #0f4a48;
            transform: translateY(-1px);
            box-shadow: 0 4px 8px rgba(24, 95, 93, 0.3);
        }
        .stFileUploader > div {
            border: 2px dashed #185f5d;
            border-radius: 8px;
            background-color: #f8f9fa;
        }
        .stFileUploader > div:hover {
            border-color: #0f4a48;
            background-color: #e9ecef;
        }
        .stDataFrame {
            border: 1px solid #185f5d;
            border-radius: 8px;
        }
        .stDataFrame th {
            background-color: #185f5d !important;
            color: white !important;
        }
        .stDataFrame td {
            border-bottom: 1px solid #e9ecef;
        }
        .stDataFrame tr:nth-child(even) {
            background-color: #f8f9fa;
        }
        .stDataFrame tr:hover {
            background-color: #e9ecef;
        }
        .stAlert {
            border-left: 4px solid #185f5d;
            background-color: #e9ecef;
        }
        .stAlert [data-testid="stMarkdown"] {
            color: #000000;
        }
        .stInfo {
            background-color: #d1ecf1;
            border: 1px solid #bee5eb;
            color: #0c5460;
        }
        .stError {
            background-color: #f8d7da;
            border: 1px solid #f5c6cb;
            color: #721c24;
        }
        .stSuccess {
            background-color: #d4edda;
            border: 1px solid #c3e6cb;
            color: #155724;
        }
        .stNumberInput > div {
            max-width: 300px !important;
        }
        .stNumberInput > div > div {
            max-width: 300px !important;
        }
        .stNumberInput input {
            max-width: 300px !important;
        }
        .stButton > button:disabled {
            background-color: #e9ecef !important;
            color: #6c757d !important;
            border: 1px solid #dee2e6 !important;
            cursor: not-allowed !important;
            opacity: 0.6 !important;
        }
        .stButton > button:disabled:hover {
            background-color: #e9ecef !important;
            color: #6c757d !important;
            transform: none !important;
            box-shadow: none !important;
        }
    </style>
    <div class="logo-header">
        <img src="https://www.dandpo.com/media/logo/stores/1/dandpo-logo-01.svg" alt="Logo" style="height: 60px;">
    </div>
    <div class="title-section">
        <h1>Supplier Sheet Builder</h1>
        <div class="separator-line"></div>
    </div>
""", unsafe_allow_html=True)
tabs = st.tabs(["Bidding Sheet Builder", "Catalog Sheet Builder"])

with tabs[0]:
    st.header("📊 Bidding Sheet Builder")
    REQUIRED_COLUMNS = [
        'Category',
        'Dandpo SKU',
        'Combinations',
        'Printer Specifications',
        'Quantity',
        'Sample',
        'Printer Cost',
        'Lead Time',
        'Weight in kg',
        'Partner Name'
    ]

    KEY_COLUMNS = [
        'Category',
        'Dandpo SKU',
        'Combinations',
        'Printer Specifications',
        'Quantity',
        'Sample',
        'Lead Time',
        'Weight in kg'
    ]

    PRICE_COLUMN = 'Printer Cost'
    PARTNER_COLUMN = 'Partner Name'

    def validate_columns(df, filename):
        missing = [col for col in REQUIRED_COLUMNS if col not in df.columns]
        if missing:
            raise ValueError(f"File {filename} is missing columns: {', '.join(missing)}")
        return df

    def generate_colored_excel(df):
        wb = Workbook()
        ws = wb.active
        ws.title = "Bidding Sheet"
        green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            ws.append(row)
            if r_idx == 1:
                continue
            bid_price = row[df.columns.get_loc("Bid Selected Price")]
            for col_idx, cell_val in enumerate(row, 1):
                if isinstance(cell_val, (int, float)) and isinstance(bid_price, (int, float)) and round(cell_val, 2) == round(bid_price, 2):
                    ws.cell(row=r_idx, column=col_idx).fill = green_fill
        bio = io.BytesIO()
        wb.save(bio)
        bio.seek(0)
        return bio

    def style_dataframe(df):
        def highlight_min_price(row):
            bid_price = row.get("Bid Selected Price")
            return [
                "background-color: #90EE90" if isinstance(col_val, (int, float)) and round(col_val, 2) == round(bid_price, 2) else ""
                for col_val in row
            ]

        def format_cell(col):
            def formatter(x):
                if pd.isna(x):
                    return ""
                try:
                    num = float(x)
                    return f"{int(num)}" if num.is_integer() else f"{num:.2f}"
                except:
                    return str(x)
            return formatter

        formatters = {col: format_cell(col) for col in df.columns}
        return df.style.apply(highlight_min_price, axis=1).format(formatters, na_rep="")

    uploaded_files = st.file_uploader(
        "Upload Supplier Cost Sheets (CSV format)", type=["csv"], accept_multiple_files=True, key="bidding_uploader"
    )
    
    # Simple sample file download
    sample_data = {
        'Category': ['T-Shirts', 'T-Shirts', 'Hoodies', 'Hoodies'],
        'Dandpo SKU': ['TS001', 'TS001', 'HD001', 'HD001'],
        'Combinations': ['Red, L', 'Red, XL', 'Blue, M', 'Blue, L'],
        'Printer Specifications': ['DTG Print', 'DTG Print', 'Screen Print', 'Screen Print'],
        'Quantity': [1, 10, 75, 100],
        'Sample': ['Yes', 'No', 'No', 'No'],
        'Printer Cost': [30, 250, 1500, 1800],
        'Lead Time': [48, 48, 74, 74],
        'Weight in kg': [0.2, 0.2, 1.4, 2.4],
        'Partner Name': ['Pname', 'Pname', 'Pname', 'Pname']
    }
    sample_df = pd.DataFrame(sample_data)
    sample_csv = sample_df.to_csv(index=False).encode("utf-8")
    st.download_button(
        "📥 Download sample file",
        sample_csv,
        file_name="Supplier_Cost_Sheet_Sample.csv",
        mime="text/csv"
    )

    if uploaded_files:
        # Initialize session state for tracking applied percentage
        if 'bidding_applied_percentage' not in st.session_state:
            st.session_state.bidding_applied_percentage = 35.0
            
        # Percentage adjustment for Bidding Sheet
        markup_percentage = st.number_input(
            "Customer Price Markup (%)",
            min_value=0.0,
            max_value=200.0,
            value=st.session_state.bidding_applied_percentage,
            step=0.1,
            help="Enter the percentage markup to apply to the bid selected price (press Enter to apply)",
            key="bidding_markup"
        )
        
        # Update applied percentage when input changes
        if 'bidding_previous_value' not in st.session_state:
            st.session_state.bidding_previous_value = st.session_state.bidding_applied_percentage
            
        # Check if value changed or if user pressed Enter (same value but different from previous)
        if (markup_percentage != st.session_state.bidding_applied_percentage or 
            (markup_percentage == st.session_state.bidding_applied_percentage and 
             markup_percentage != st.session_state.bidding_previous_value)):
            st.session_state.bidding_applied_percentage = markup_percentage
            
        st.session_state.bidding_previous_value = markup_percentage
            
        st.write(f"**Current markup:** {st.session_state.bidding_applied_percentage}% (Customer Price = Bid Price × {1 + st.session_state.bidding_applied_percentage/100:.2f})")
        # Show loader for data processing
        with st.spinner("🔄 Processing supplier data..."):
            all_data = []
            for file in uploaded_files:
                try:
                    df = pd.read_csv(file)
                    # Clean empty rows and NaN values
                    df = df.dropna(how='all')  # Remove completely empty rows
                    df = df.fillna('')  # Replace NaN with empty string
                    
                    # Convert numeric columns to proper data types
                    numeric_columns = ['Quantity', 'Printer Cost', 'Lead Time', 'Weight in kg']
                    for col in numeric_columns:
                        if col in df.columns:
                            df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)
                    
                    validate_columns(df, file.name)
                    all_data.append(df)
                except Exception as e:
                    st.error(f"❌ {file.name}: {e}")
                    st.stop()
            combined_df = pd.concat(all_data, ignore_index=True)
        combined_df['Product Key'] = combined_df[KEY_COLUMNS].astype(str).agg(' | '.join, axis=1)
        grouped = combined_df.groupby('Product Key')
        output_rows = []
        all_partners = combined_df[PARTNER_COLUMN].unique().tolist()
        for key, group in grouped:
            base_data = group.iloc[0][KEY_COLUMNS].to_dict()
            quantity = group.iloc[0]['Quantity']
            partner_prices = {}
            for _, row in group.iterrows():
                partner = row[PARTNER_COLUMN]
                price = row[PRICE_COLUMN]
                # Only consider prices > 0 (participating suppliers)
                if pd.notna(price) and price > 0:
                    partner_prices[partner] = price
            if partner_prices:
                min_price = min(partner_prices.values())
                winners = [k for k, v in partner_prices.items() if v == min_price]
                winner_str = ", ".join(winners)
            else:
                min_price = None
                winner_str = None
            # Calculate customer price using the applied percentage
            customer_price = int(round(min_price * (1 + st.session_state.bidding_applied_percentage/100))) if min_price is not None else None
            unit_price = min_price / quantity if min_price is not None and quantity else None
            customer_unit_price = customer_price / quantity if customer_price is not None and quantity else None

            output_row = {
                **base_data,
                'Bid Selected Partners': winner_str,
                'Bid Selected Price': min_price,
                f'Customer Price ({st.session_state.bidding_applied_percentage}%)': customer_price,
                'Bid Selected Unit Price': unit_price,
                'Customer Unit Price': customer_unit_price
            }
            for partner in all_partners:
                output_row[partner] = partner_prices.get(partner, None)
            output_rows.append(output_row)

        bidding_df = pd.DataFrame(output_rows)
        st.subheader("📄 Bidding Sheet Preview")
        st.dataframe(style_dataframe(bidding_df), use_container_width=True)

        csv_bytes = bidding_df.to_csv(index=False).encode("utf-8")
        st.download_button(
            "📥 Download as CSV",
            csv_bytes,
            file_name="Bidding Sheet.csv",
            mime="text/csv"
        )
        excel_file = generate_colored_excel(bidding_df)
        st.download_button(
            "📥 Download as Excel (Colored)",
            excel_file,
            file_name="Bidding Sheet.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    else:
        st.info("Please upload at least one supplier cost sheet in CSV format.")

with tabs[1]:
    st.header("📘 Catalog Sheet Builder")
    
    catalog_upload = st.file_uploader("Upload Bidding Sheet CSV", type=["csv"], key="catalog_uploader")
    
    # Simple sample file download
    sample_bidding_data = {
        'Category': ['T-Shirts', 'Hoodies', 'Caps'],
        'Dandpo SKU': ['TS001', 'HD001', 'CP001'],
        'Combinations': ['Red, L', 'Blue, M', 'Black, One Size'],
        'Printer Specifications': ['DTG Print', 'Screen Print', 'Embroidery'],
        'Quantity': [100, 75, 50],
        'Sample': ['Yes', 'No', 'Yes'],
        'Lead Time': [7, 10, 5],
        'Weight in kg': [0.2, 0.4, 0.1],
        'Bid Selected Partners': ['Pname1', 'Pname1', 'Pname1'],
        'Bid Selected Price': [550, 618, 275],
        'Customer Price (35%)': [743, 834, 371],
        'Bid Selected Unit Price': [5.50, 8.24, 5.50],
        'Customer Unit Price': [7.43, 11.12, 7.42],
        'Pname1': [550, 618, 275],
        'Pname2': [575, 650, 290]
    }
    sample_bidding_df = pd.DataFrame(sample_bidding_data)
    sample_bidding_csv = sample_bidding_df.to_csv(index=False).encode("utf-8")
    st.download_button(
        "📥 Download sample file",
        sample_bidding_csv,
        file_name="Bidding_Sheet_Sample.csv",
        mime="text/csv"
    )

    if catalog_upload:
        bidding_df = pd.read_csv(catalog_upload)
        # Clean empty rows and NaN values
        bidding_df = bidding_df.dropna(how='all')  # Remove completely empty rows
        bidding_df = bidding_df.fillna('')  # Replace NaN with empty string
        
        # Convert numeric columns to proper data types
        numeric_columns = ['Quantity', 'Bid Selected Price', 'Lead Time', 'Weight in kg', 'Bid Selected Unit Price', 'Customer Unit Price']
        for col in numeric_columns:
            if col in bidding_df.columns:
                bidding_df[col] = pd.to_numeric(bidding_df[col], errors='coerce').fillna(0)
        
        # Convert customer price columns to numeric
        customer_price_cols = [col for col in bidding_df.columns if 'Customer Price' in col]
        for col in customer_price_cols:
            if col in bidding_df.columns:
                bidding_df[col] = pd.to_numeric(bidding_df[col], errors='coerce').fillna(0)
        
        # Initialize session state for tracking applied percentage
        if 'catalog_applied_percentage' not in st.session_state:
            st.session_state.catalog_applied_percentage = 0.0
            
        # Percentage adjustment for Catalog Sheet
        catalog_markup_percentage = st.number_input(
            "Recalculate Customer Price Markup (%)",
            min_value=0.0,
            max_value=200.0,
            value=st.session_state.catalog_applied_percentage,
            step=0.1,
            help="Enter the percentage markup to recalculate customer prices from printer costs (press Enter to apply)",
            key="catalog_markup"
        )
        
        # Update applied percentage when input changes
        if 'catalog_previous_value' not in st.session_state:
            st.session_state.catalog_previous_value = st.session_state.catalog_applied_percentage
            
        # Check if value changed or if user pressed Enter (same value but different from previous)
        if (catalog_markup_percentage != st.session_state.catalog_applied_percentage or 
            (catalog_markup_percentage == st.session_state.catalog_applied_percentage and 
             catalog_markup_percentage != st.session_state.catalog_previous_value)):
            st.session_state.catalog_applied_percentage = catalog_markup_percentage
            
        st.session_state.catalog_previous_value = catalog_markup_percentage
        
        catalog_cols = [
            'Category', 'Dandpo SKU', 'Combinations', 'Printer Specifications',
            'Quantity', 'Sample', 'Lead Time', 'Weight in kg', 'Printer Cost',
            'Customer Price', 'Unit Price for Printer cost', 'Unit Price for customer cost',
            'Partners', 'Production Type', 'Production Time (Hours)',
            'Printer Delivery to Dandpo (Hours)', 'Packaging Dimensions', 'Product Dimensions'
        ]

        catalog_df = pd.DataFrame()
        catalog_df['Category'] = bidding_df['Category']
        catalog_df['Dandpo SKU'] = bidding_df['Dandpo SKU']
        catalog_df['Combinations'] = bidding_df['Combinations']
        catalog_df['Printer Specifications'] = bidding_df['Printer Specifications']
        catalog_df['Quantity'] = bidding_df['Quantity']
        catalog_df['Sample'] = bidding_df['Sample']
        catalog_df['Lead Time'] = bidding_df['Lead Time']
        catalog_df['Weight in kg'] = bidding_df['Weight in kg']
        catalog_df['Printer Cost'] = bidding_df['Bid Selected Price']

        quantity = bidding_df['Quantity'].replace(0, pd.NA).astype(float)
        printer_cost = bidding_df['Bid Selected Price'].astype(float)

        # Initially copy customer price from bidding sheet
        customer_price_cols = [col for col in bidding_df.columns if 'Customer Price' in col]
        if customer_price_cols:
            # Copy the exact column name and values from bidding sheet
            customer_price_col = customer_price_cols[0]
            catalog_df[customer_price_col] = bidding_df[customer_price_col]
        else:
            # Fallback if no customer price column exists
            catalog_df['Customer Price'] = printer_cost
        catalog_df['Unit Price for Printer cost'] = bidding_df.get('Bid Selected Unit Price', printer_cost / quantity).round(2)
        # Calculate unit price for customer cost using the correct customer price column
        customer_price_cols = [col for col in bidding_df.columns if 'Customer Price' in col]
        if customer_price_cols:
            customer_price_col = customer_price_cols[0]
            catalog_df['Unit Price for customer cost'] = bidding_df.get('Customer Unit Price', catalog_df[customer_price_col] / quantity).round(2)
        else:
            # Use the fallback customer price column we created
            catalog_df['Unit Price for customer cost'] = bidding_df.get('Customer Unit Price', catalog_df['Customer Price'] / quantity).round(2)

        # Recalculate customer prices if percentage is not 0
        if st.session_state.catalog_applied_percentage > 0:
            # Find the existing customer price column and replace its values
            customer_price_cols = [col for col in catalog_df.columns if 'Customer Price' in col]
            if customer_price_cols:
                existing_col = customer_price_cols[0]
                # Replace the existing column with new values
                new_customer_price = round(printer_cost * (1 + st.session_state.catalog_applied_percentage/100)).round().astype('Int64')
                catalog_df[existing_col] = new_customer_price
                # Rename the column to show the new percentage
                catalog_df = catalog_df.rename(columns={existing_col: f'Customer Price ({st.session_state.catalog_applied_percentage}%)'})
                # Update unit price for customer cost
                catalog_df['Unit Price for customer cost'] = (new_customer_price / quantity).round(2)
            else:
                # Create new column if none exists
                new_customer_price = round(printer_cost * (1 + st.session_state.catalog_applied_percentage/100)).round().astype('Int64')
                catalog_df[f'Customer Price ({st.session_state.catalog_applied_percentage}%)'] = new_customer_price
                # Update unit price for customer cost
                catalog_df['Unit Price for customer cost'] = (new_customer_price / quantity).round(2)

        # Exclude customer price columns from partners calculation
        customer_price_cols = [col for col in bidding_df.columns if 'Customer Price' in col]
        partner_cols = [col for col in bidding_df.columns if col not in catalog_cols and col not in KEY_COLUMNS and col not in [
            'Bid Selected Partners', 'Bid Selected Price', 'Bid Selected Unit Price', 'Customer Unit Price'] + customer_price_cols]
        catalog_df['Partners'] = bidding_df[partner_cols].apply(lambda row: ', '.join(row.dropna().index), axis=1)

        # Add missing columns from catalog_cols, but skip Customer Price if we already have a customer price column
        customer_price_exists = any('Customer Price' in col for col in catalog_df.columns)
        for col in catalog_cols:
            if col not in catalog_df.columns:
                if col == 'Customer Price' and customer_price_exists:
                    continue  # Skip adding 'Customer Price' if we already have a customer price column
                catalog_df[col] = ""

        st.subheader("📘 Catalog Sheet Preview")
        st.dataframe(catalog_df, use_container_width=True)

        catalog_csv = catalog_df.to_csv(index=False).encode("utf-8")
        st.download_button(
            "📥 Download Catalog Sheet (CSV)",
            catalog_csv,
            file_name="Catalog Sheet.csv",
            mime="text/csv"
        )