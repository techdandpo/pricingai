import streamlit as st
import pandas as pd
import io
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

# Required columns for bidding sheet
REQUIRED_COLUMNS = [
    'Category',
    'Dandpo SKU',
    'Combinations',
    'Printer Specifications',
    'Quantity',
    'Sample',
    'Printer Cost',
    'Lead Time',
    'Weight in kg',
    'Partner Name'
]

KEY_COLUMNS = [
    'Category',
    'Dandpo SKU',
    'Combinations',
    'Printer Specifications',
    'Quantity',
    'Sample',
    'Lead Time',
    'Weight in kg'
]

PRICE_COLUMN = 'Printer Cost'
PARTNER_COLUMN = 'Partner Name'

def validate_columns(df, filename):
    missing = [col for col in REQUIRED_COLUMNS if col not in df.columns]
    if missing:
        raise ValueError(f"File {filename} is missing columns: {', '.join(missing)}")
    return df

def generate_colored_excel(df):
    wb = Workbook()
    ws = wb.active
    ws.title = "Bidding Sheet"
    green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        ws.append(row)
        if r_idx == 1:
            continue
        bid_price = row[df.columns.get_loc("Bid Selected Price")]
        for col_idx, cell_val in enumerate(row, 1):
            if isinstance(cell_val, (int, float)) and isinstance(bid_price, (int, float)) and round(cell_val, 2) == round(bid_price, 2):
                ws.cell(row=r_idx, column=col_idx).fill = green_fill
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio

def style_dataframe(df):
    def highlight_min_price(row):
        bid_price = row.get("Bid Selected Price")
        return [
            "background-color: #90EE90" if isinstance(col_val, (int, float)) and round(col_val, 2) == round(bid_price, 2) else ""
            for col_val in row
        ]

    def format_cell(col):
        def formatter(x):
            if pd.isna(x):
                return ""
            try:
                num = float(x)
                return f"{int(num)}" if num.is_integer() else f"{num:.2f}"
            except:
                return str(x)
        return formatter

    formatters = {col: format_cell(col) for col in df.columns}
    return df.style.apply(highlight_min_price, axis=1).format(formatters, na_rep="")

def bidding_sheet_builder():
    """Main function for Bidding Sheet Builder tab"""
    
    uploaded_files = st.file_uploader(
        "Upload Supplier Cost Sheets (CSV format)", type=["csv"], accept_multiple_files=True, key="bidding_uploader"
    )
    
    # Simple sample file download
    sample_data = {
        'Category': ['T-Shirts', 'T-Shirts', 'Hoodies', 'Hoodies'],
        'Dandpo SKU': ['TS001', 'TS001', 'HD001', 'HD001'],
        'Combinations': ['Red, L', 'Red, XL', 'Blue, M', 'Blue, L'],
        'Printer Specifications': ['DTG Print', 'DTG Print', 'Screen Print', 'Screen Print'],
        'Quantity': [1, 10, 75, 100],
        'Sample': ['Yes', 'No', 'No', 'No'],
        'Printer Cost': [30, 250, 1500, 1800],
        'Lead Time': [48, 48, 74, 74],
        'Weight in kg': [0.2, 0.2, 1.4, 2.4],
        'Partner Name': ['Pname', 'Pname', 'Pname', 'Pname']
    }
    sample_df = pd.DataFrame(sample_data)
    sample_csv = sample_df.to_csv(index=False).encode("utf-8")
    st.download_button(
        "📥 Download sample file",
        sample_csv,
        file_name="Supplier_Cost_Sheet_Sample.csv",
        mime="text/csv"
    )
    
    if uploaded_files:
        # Initialize session state for tracking applied percentage
        if 'bidding_applied_percentage' not in st.session_state:
            st.session_state.bidding_applied_percentage = 35.0
            
        # Percentage adjustment for Bidding Sheet
        markup_percentage = st.number_input(
            "Customer Price Markup (%)",
            min_value=0.0,
            max_value=200.0,
            value=st.session_state.bidding_applied_percentage,
            step=0.1,
            help="Enter the percentage markup to apply to the bid selected price (press Enter to apply)",
            key="bidding_markup"
        )
        
        # Update applied percentage when input changes
        if 'bidding_previous_value' not in st.session_state:
            st.session_state.bidding_previous_value = st.session_state.bidding_applied_percentage
            
        # Check if value changed or if user pressed Enter (same value but different from previous)
        if (markup_percentage != st.session_state.bidding_applied_percentage or 
            (markup_percentage == st.session_state.bidding_applied_percentage and 
             markup_percentage != st.session_state.bidding_previous_value)):
            st.session_state.bidding_applied_percentage = markup_percentage
            
        st.session_state.bidding_previous_value = markup_percentage
            
        st.write(f"**Current markup:** {st.session_state.bidding_applied_percentage}% (Customer Price = Bid Price × {1 + st.session_state.bidding_applied_percentage/100:.2f})")
        # Show loader for data processing
        with st.spinner("🔄 Processing supplier data..."):
            all_data = []
            for file in uploaded_files:
                try:
                    df = pd.read_csv(file)
                    # Clean empty rows and NaN values
                    df = df.dropna(how='all')  # Remove completely empty rows
                    df = df.fillna('')  # Replace NaN with empty string
                    
                    # Convert numeric columns to proper data types
                    numeric_columns = ['Quantity', 'Printer Cost', 'Lead Time', 'Weight in kg']
                    for col in numeric_columns:
                        if col in df.columns:
                            df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)
                    
                    validate_columns(df, file.name)
                    all_data.append(df)
                except Exception as e:
                    st.error(f"❌ {file.name}: {e}")
                    st.stop()
            combined_df = pd.concat(all_data, ignore_index=True)
        combined_df['Product Key'] = combined_df[KEY_COLUMNS].astype(str).agg(' | '.join, axis=1)
        grouped = combined_df.groupby('Product Key')
        output_rows = []
        all_partners = combined_df[PARTNER_COLUMN].unique().tolist()
        for key, group in grouped:
            base_data = group.iloc[0][KEY_COLUMNS].to_dict()
            quantity = group.iloc[0]['Quantity']
            partner_prices = {}
            for _, row in group.iterrows():
                partner = row[PARTNER_COLUMN]
                price = row[PRICE_COLUMN]
                # Only consider prices > 0 (participating suppliers)
                if pd.notna(price) and price > 0:
                    partner_prices[partner] = price
            if partner_prices:
                min_price = min(partner_prices.values())
                winners = [k for k, v in partner_prices.items() if v == min_price]
                winner_str = ", ".join(winners)
            else:
                min_price = None
                winner_str = None
            # Calculate customer price using the applied percentage
            customer_price = int(round(min_price * (1 + st.session_state.bidding_applied_percentage/100))) if min_price is not None else None
            unit_price = min_price / quantity if min_price is not None and quantity else None
            customer_unit_price = customer_price / quantity if customer_price is not None and quantity else None

            output_row = {
                **base_data,
                'Bid Selected Partners': winner_str,
                'Bid Selected Price': min_price,
                f'Customer Price ({st.session_state.bidding_applied_percentage}%)': customer_price,
                'Bid Selected Unit Price': unit_price,
                'Customer Unit Price': customer_unit_price
            }
            for partner in all_partners:
                output_row[partner] = partner_prices.get(partner, None)
            output_rows.append(output_row)

        bidding_df = pd.DataFrame(output_rows)
        st.subheader("📄 Bidding Sheet Preview")
        st.dataframe(style_dataframe(bidding_df), use_container_width=True)

        csv_bytes = bidding_df.to_csv(index=False).encode("utf-8")
        st.download_button(
            "📥 Download as CSV",
            csv_bytes,
            file_name="Bidding Sheet.csv",
            mime="text/csv"
        )
        excel_file = generate_colored_excel(bidding_df)
        st.download_button(
            "📥 Download as Excel (Colored)",
            excel_file,
            file_name="Bidding Sheet.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    else:
        st.info("Please upload at least one supplier cost sheet in CSV format.")
